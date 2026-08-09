#!/usr/bin/env python3
"""Prove the image can read a real spreadsheet, at build time.

The agent's first XLSX import raised `ModuleNotFoundError: openpyxl`. It
recovered by unzipping the workbook and reading the XML itself, which works
for a simple sheet and quietly misreads the things that make a real one hard:
a merged header, a formula cell, a second sheet, non-ASCII text, a trailing
total row that is not data.

So the fixture here is built out of exactly those, parsed back, and asserted.
Running it as a build step means a fresh Pod cannot start without the
toolchain — the failure lands on whoever is building, not on an agent halfway
through someone's import.
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ModuleNotFoundError:  # pragma: no cover - this is the failure being prevented
    print("openpyxl is missing: an agent asked to read an .xlsx would fall back to raw XML", file=sys.stderr)
    raise


def build(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品清单"

    # a merged banner above the header row — the shape that makes a naive
    # reader treat row 1 as the header and lose every column name
    sheet["A1"] = "2026 年度产品目录"
    sheet.merge_cells("A1:D1")
    sheet.append([])
    sheet.append(["物料号", "品名", "单价", "数量"])
    sheet.append(["P-001", "内窥镜镜头", 1200.50, 3])
    sheet.append(["P-002", "光源模块", 8600.00, 2])
    # a formula, not a literal
    sheet["E4"] = "=C4*D4"
    # a trailing total row: present in the sheet, never a product
    sheet.append(["合计", None, None, 5])

    second = workbook.create_sheet("供应商")
    second.append(["供应商编码", "名称"])
    second.append(["V-001", "华东医疗器械"])

    workbook.save(path)


def check(path: Path) -> None:
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["产品清单", "供应商"], workbook.sheetnames

    sheet = workbook["产品清单"]
    assert {str(r) for r in sheet.merged_cells.ranges} == {"A1:D1"}, sheet.merged_cells.ranges
    header = [cell.value for cell in sheet[3]]
    assert header[:4] == ["物料号", "品名", "单价", "数量"], header
    assert sheet["B4"].value == "内窥镜镜头", sheet["B4"].value
    assert sheet["C4"].value == 1200.50, sheet["C4"].value
    # row 6, not 7: setting E4 directly does not extend max_row, so the total
    # lands right after the last appended product
    assert sheet["A6"].value == "合计", sheet["A6"].value
    assert sheet.max_row == 6, sheet.max_row

    # the formula is readable as a formula; a values-only read gives the cached
    # result, which is None until Excel has computed it — an importer that does
    # not know the difference reports a blank price
    assert sheet["E4"].value == "=C4*D4", sheet["E4"].value
    cached = load_workbook(path, data_only=True)["产品清单"]["E4"].value
    assert cached is None, f"expected no cached value in a file openpyxl wrote, got {cached!r}"

    assert workbook["供应商"]["B2"].value == "华东医疗器械"


if __name__ == "__main__":
    with tempfile.TemporaryDirectory() as directory:
        fixture = Path(directory) / "fixture.xlsx"
        build(fixture)
        check(fixture)
    print("XLSX TOOLCHAIN OK")
